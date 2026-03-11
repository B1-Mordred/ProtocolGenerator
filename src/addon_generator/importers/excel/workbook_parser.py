from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from addon_generator.importers.excel_importer import ExcelImportValidationError, ImportDiagnostic
from addon_generator.importers.excel.analytes_parser import parse_analytes_sheet
from addon_generator.importers.excel.basics_parser import parse_basics_sheet
from addon_generator.importers.excel.dilutions_parser import parse_dilutions_sheet
from addon_generator.importers.excel.sampleprep_parser import parse_sampleprep_sheet
from addon_generator.input_models.dtos import InputDTOBundle


READ_ONLY_SHEETS = {"AddOn CheckList"}
VOCAB_SHEETS = {"Hidden_Lists"}


@dataclass(slots=True)
class WorkbookParseArtifacts:
    bundle: InputDTOBundle
    diagnostics: list[ImportDiagnostic]


class ExcelWorkbookParser:
    """Parses workbook-style addon templates into DTOs."""

    def parse_path(self, excel_path: str | Path) -> InputDTOBundle:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils.exceptions import InvalidFileException  # type: ignore

        workbook_path = Path(excel_path)
        try:
            wb = load_workbook(workbook_path, data_only=True)
        except Exception as exc:
            rule_id = "invalid-workbook-format" if isinstance(exc, (BadZipFile, InvalidFileException)) else "workbook-open-failed"
            raise ExcelImportValidationError(
                "Workbook could not be opened",
                [
                    ImportDiagnostic(
                        rule_id=rule_id,
                        message="Workbook payload is not a readable .xlsx archive",
                        sheet="(workbook)",
                        value={
                            "path": str(workbook_path),
                            "error_type": type(exc).__name__,
                            "error_message": str(exc),
                        },
                    )
                ],
            ) from exc
        artifacts = self.parse_workbook(wb, source_name=str(workbook_path))
        if artifacts.diagnostics:
            raise ExcelImportValidationError("Workbook contains validation errors", artifacts.diagnostics)
        return artifacts.bundle

    def parse_workbook(self, workbook: Any, *, source_name: str | None = None) -> WorkbookParseArtifacts:
        diagnostics: list[ImportDiagnostic] = []
        sheet_names = self._sheet_lookup(workbook.sheetnames)

        vocab = self._extract_vocab(workbook[sheet_names["hiddenlists"]]) if "hiddenlists" in sheet_names else {}

        basics = parse_basics_sheet(workbook[sheet_names["basics"]], diagnostics=diagnostics) if "basics" in sheet_names else None
        assay_lookup_by_parameter_set = basics.assay_reference_lookup if basics is not None else {}
        analytes = (
            parse_analytes_sheet(
                workbook[sheet_names["analytes"]],
                vocab=vocab,
                diagnostics=diagnostics,
                assay_lookup_by_parameter_set=assay_lookup_by_parameter_set,
            )
            if "analytes" in sheet_names
            else None
        )
        sample_prep = parse_sampleprep_sheet(workbook[sheet_names["sampleprep"]], vocab=vocab, diagnostics=diagnostics) if "sampleprep" in sheet_names else []
        dilutions = parse_dilutions_sheet(workbook[sheet_names["dilutions"]], diagnostics=diagnostics) if "dilutions" in sheet_names else []

        if basics is None:
            diagnostics.append(ImportDiagnostic(rule_id="missing-sheet", message="Required Basics sheet is missing", sheet="(workbook)"))
        if analytes is None:
            diagnostics.append(ImportDiagnostic(rule_id="missing-sheet", message="Required Analytes sheet is missing", sheet="(workbook)"))

        bundle = InputDTOBundle(source_type="excel", source_name=source_name)
        if basics is not None:
            bundle.method = basics.method
            bundle.assays = basics.assays
        if analytes is not None:
            bundle.analytes = analytes.analytes
            bundle.units = analytes.units
        bundle.sample_prep_steps = sample_prep
        bundle.dilution_schemes = dilutions
        bundle.hidden_vocab = {key: sorted(values) for key, values in vocab.items()}

        return WorkbookParseArtifacts(bundle=bundle, diagnostics=diagnostics)

    def _extract_vocab(self, sheet: Any) -> dict[str, set[str]]:
        header = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header:
            return {}
        headers = [self._text(cell.value) for cell in header]
        vocab: dict[str, set[str]] = {h: set() for h in headers if h}
        for row in sheet.iter_rows(min_row=2):
            for idx, name in enumerate(headers):
                if not name:
                    continue
                value = self._text(row[idx].value)
                if value:
                    vocab[name].add(value)
        return vocab

    @staticmethod
    def supports_workbook_template(sheet_names: list[str]) -> bool:
        effective_sheets = {
            ExcelWorkbookParser._normalize_sheet_name(name)
            for name in sheet_names
            if ExcelWorkbookParser._normalize_sheet_name(name)
            not in {
                ExcelWorkbookParser._normalize_sheet_name(value)
                for value in READ_ONLY_SHEETS | VOCAB_SHEETS
            }
        }
        return {"basics", "analytes"}.issubset(effective_sheets)

    @staticmethod
    def _normalize_sheet_name(name: str) -> str:
        return str(name or "").strip().replace(" ", "").replace("_", "").casefold()

    def _sheet_lookup(self, sheet_names: list[str]) -> dict[str, str]:
        return {self._normalize_sheet_name(name): name for name in sheet_names}

    @staticmethod
    def _text(value: Any) -> str:
        if value is None:
            return ""
        return str(value).strip()

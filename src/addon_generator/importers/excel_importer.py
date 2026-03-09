from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any
from zipfile import BadZipFile

from addon_generator.domain.models import AddonModel, normalize_assay_identity_fields
from addon_generator.importers.gui_mapper import map_gui_payload_to_bundle
from addon_generator.input_models.dtos import InputDTOBundle
from addon_generator.input_models.provenance import FieldProvenance
from addon_generator.services.canonical_model_builder import CanonicalModelBuilder


@dataclass(frozen=True, slots=True)
class ImportDiagnostic:
    rule_id: str
    message: str
    sheet: str
    row: int | None = None
    column: str | None = None
    value: Any | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            "rule_id": self.rule_id,
            "message": self.message,
            "sheet": self.sheet,
            "row": self.row,
            "column": self.column,
            "value": self.value,
        }


class ExcelImportValidationError(ValueError):
    def __init__(self, message: str, diagnostics: list[ImportDiagnostic]):
        super().__init__(message)
        self.diagnostics = diagnostics

    def to_dict(self) -> dict[str, Any]:
        return {"message": str(self), "diagnostics": [d.to_dict() for d in self.diagnostics]}


class ExcelImporter:
    COLUMN_MAPPINGS: dict[str, dict[str, dict[str, str]]] = {
        "v1-flat": {
            "*": {
                "MethodId": "method_id",
                "MethodVersion": "method_version",
                "MethodDisplayName": "method_display_name",
                "AssayKey": "assay_key",
                "ProtocolType": "protocol_type",
                "AssayDisplayName": "assay_display_name",
                "XmlAssayName": "xml_assay_name",
                "AnalyteKey": "analyte_key",
                "AnalyteName": "analyte_name",
                "AssayInformationType": "assay_information_type",
                "UnitKey": "unit_key",
                "UnitName": "unit_name",
            }
        },
        "v2-sheeted": {
            "Method": {"MethodId": "method_id", "MethodVersion": "method_version", "MethodDisplayName": "method_display_name"},
            "Assays": {
                "AssayKey": "assay_key",
                "ProtocolType": "protocol_type",
                "AssayDisplayName": "assay_display_name",
                "XmlAssayName": "xml_assay_name",
            },
            "Analytes": {
                "AnalyteKey": "analyte_key",
                "AnalyteName": "analyte_name",
                "AssayKey": "assay_key",
                "AssayInformationType": "assay_information_type",
            },
            "Units": {"UnitKey": "unit_key", "UnitName": "unit_name", "AnalyteKey": "analyte_key"},
        },
    }

    REQUIRED_COLUMNS: dict[str, dict[str, set[str]]] = {
        "v1-flat": {"*": {"MethodId", "MethodVersion", "AssayKey", "ProtocolType", "AnalyteKey", "AnalyteName", "UnitKey", "UnitName"}},
        "v2-sheeted": {
            "Method": {"MethodId", "MethodVersion"},
            "Assays": {"AssayKey", "ProtocolType"},
            "Analytes": {"AnalyteKey", "AnalyteName", "AssayKey"},
            "Units": {"UnitKey", "UnitName", "AnalyteKey"},
        },
    }

    def read_workbook_rows(self, excel_path: str | Path) -> list[dict[str, Any]]:
        workbook_data = self._parse_workbook_rows(excel_path)
        if workbook_data["layout_version"] == "v1-flat":
            return workbook_data["rows"]
        rows: list[dict[str, Any]] = []
        method = workbook_data["method"]
        assays = workbook_data["assays"]
        analytes = workbook_data["analytes"]
        units = workbook_data["units"]

        assays_by_key: dict[str, dict[str, Any]] = {
            self._to_string(assay.get("assay_key")): assay
            for assay in assays
            if self._to_string(assay.get("assay_key"))
        }
        analytes_by_key: dict[str, dict[str, Any]] = {
            self._to_string(analyte.get("analyte_key")): analyte
            for analyte in analytes
            if self._to_string(analyte.get("analyte_key"))
        }

        def _base_row() -> dict[str, Any]:
            return {
                "MethodId": method.get("method_id"),
                "MethodVersion": method.get("method_version"),
                "MethodDisplayName": method.get("method_display_name"),
                "AssayKey": None,
                "ProtocolType": None,
                "AssayDisplayName": None,
                "XmlAssayName": None,
                "AnalyteKey": None,
                "AnalyteName": None,
                "AssayInformationType": None,
                "UnitKey": None,
                "UnitName": None,
            }

        for assay in assays:
            row = _base_row()
            row.update(
                {
                    "AssayKey": assay.get("assay_key"),
                    "ProtocolType": assay.get("protocol_type"),
                    "AssayDisplayName": assay.get("assay_display_name"),
                    "XmlAssayName": assay.get("xml_assay_name"),
                }
            )
            rows.append(row)

        for analyte in analytes:
            row = _base_row()
            assay = assays_by_key.get(self._to_string(analyte.get("assay_key")))
            if assay is not None:
                row.update(
                    {
                        "AssayKey": assay.get("assay_key"),
                        "ProtocolType": assay.get("protocol_type"),
                        "AssayDisplayName": assay.get("assay_display_name"),
                        "XmlAssayName": assay.get("xml_assay_name"),
                    }
                )
            else:
                row["AssayKey"] = analyte.get("assay_key")
            row.update(
                {
                    "AnalyteKey": analyte.get("analyte_key"),
                    "AnalyteName": analyte.get("analyte_name"),
                    "AssayInformationType": analyte.get("assay_information_type"),
                }
            )
            rows.append(row)

        for unit in units:
            row = _base_row()
            analyte = analytes_by_key.get(self._to_string(unit.get("analyte_key")))
            if analyte is not None:
                row.update(
                    {
                        "AnalyteKey": analyte.get("analyte_key"),
                        "AnalyteName": analyte.get("analyte_name"),
                        "AssayInformationType": analyte.get("assay_information_type"),
                    }
                )
                assay = assays_by_key.get(self._to_string(analyte.get("assay_key")))
                if assay is not None:
                    row.update(
                        {
                            "AssayKey": assay.get("assay_key"),
                            "ProtocolType": assay.get("protocol_type"),
                            "AssayDisplayName": assay.get("assay_display_name"),
                            "XmlAssayName": assay.get("xml_assay_name"),
                        }
                    )
                else:
                    row["AssayKey"] = analyte.get("assay_key")
            else:
                row["AnalyteKey"] = unit.get("analyte_key")
            row.update({"UnitKey": unit.get("unit_key"), "UnitName": unit.get("unit_name")})
            rows.append(row)
        return rows

    def normalize_workbook_rows(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        if not rows:
            return {"method_id": "", "method_version": "", "assays": [], "analytes": [], "units": []}

        first = rows[0]
        payload: dict[str, Any] = {
            "method_id": self._to_string(first.get("MethodId")),
            "method_version": self._to_string(first.get("MethodVersion")),
            "MethodInformation": {"DisplayName": self._to_string(first.get("MethodDisplayName"))},
            "assays": [],
            "analytes": [],
            "units": [],
        }
        for row in rows:
            assay_key = self._to_string(row.get("AssayKey")) or self._to_string(row.get("AssayDisplayName"))
            analyte_key = self._to_string(row.get("AnalyteKey")) or self._to_string(row.get("AnalyteName"))
            unit_raw = self._to_string(row.get("UnitName"))
            unit_names = self._split_multi_units(unit_raw)
            unit_key = self._to_string(row.get("UnitKey")) or (unit_names[0] if unit_names else "")
            if assay_key:
                protocol_type, protocol_display_name, xml_name = normalize_assay_identity_fields(
                    protocol_type=self._to_string(row.get("ProtocolType")),
                    protocol_display_name=self._to_string(row.get("AssayDisplayName")),
                    xml_name=self._to_string(row.get("XmlAssayName")),
                )
                if any((protocol_type, protocol_display_name, xml_name)):
                    payload["assays"].append(
                        {
                            "key": assay_key,
                            "protocol_type": protocol_type,
                            "protocol_display_name": protocol_display_name,
                            "xml_name": xml_name,
                        }
                    )
            analyte_name = self._to_string(row.get("AnalyteName"))
            if analyte_key and analyte_name:
                payload["analytes"].append(
                    {
                        "key": analyte_key,
                        "name": analyte_name,
                        "assay_key": assay_key,
                        "assay_information_type": self._to_string(row.get("AssayInformationType")),
                    }
                )
            if unit_names:
                for idx, unit_name in enumerate(unit_names):
                    payload["units"].append(
                        {
                            "key": unit_key if idx == 0 else f"{unit_key}:{idx}",
                            "name": self._normalize_unit_name(unit_name),
                            "analyte_key": analyte_key,
                        }
                    )
            elif unit_key:
                payload["units"].append(
                    {
                        "key": unit_key,
                        "name": self._normalize_unit_name(self._to_string(row.get("UnitName"))),
                        "analyte_key": analyte_key,
                    }
                )
        return payload

    def map_workbook_rows_to_dto_bundle(self, rows: list[dict[str, Any]], *, source_name: str | None = None) -> InputDTOBundle:
        bundle = map_gui_payload_to_bundle(self.normalize_workbook_rows(rows))
        bundle.source_type = "excel"
        bundle.source_name = source_name
        bundle.provenance.setdefault("method.method_id", []).append(
            FieldProvenance(source_type="excel", source_file=source_name, source_sheet="(workbook)", field_key="MethodId")
        )
        return bundle

    def map_workbook_rows_to_canonical_model(self, rows: list[dict[str, Any]]) -> AddonModel:
        return CanonicalModelBuilder().build(self.map_workbook_rows_to_dto_bundle(rows))

    def import_workbook_bundle(self, excel_path: str | Path) -> InputDTOBundle:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils.exceptions import InvalidFileException  # type: ignore

        from addon_generator.importers.excel.workbook_parser import ExcelWorkbookParser

        workbook_path = Path(excel_path)
        try:
            workbook = load_workbook(workbook_path, data_only=True)
        except Exception as exc:
            rule_id = "invalid-workbook-format" if isinstance(exc, (BadZipFile, InvalidFileException)) else "workbook-open-failed"
            raise ExcelImportValidationError(
                "Workbook could not be opened",
                [
                    ImportDiagnostic(
                        rule_id=rule_id,
                        message="Workbook payload is not a readable .xlsx archive",
                        sheet="(workbook)",
                        value={
                            "path": str(workbook_path),
                            "error_type": type(exc).__name__,
                            "error_message": str(exc),
                        },
                    )
                ],
            ) from exc
        if ExcelWorkbookParser.supports_workbook_template(workbook.sheetnames):
            artifacts = ExcelWorkbookParser().parse_workbook(workbook, source_name=str(excel_path))
            if artifacts.diagnostics:
                raise ExcelImportValidationError("Workbook contains validation errors", artifacts.diagnostics)
            return artifacts.bundle

        rows = self.read_workbook_rows(excel_path)
        return self.map_workbook_rows_to_dto_bundle(rows, source_name=str(excel_path))

    def import_workbook(self, excel_path: str | Path) -> AddonModel:
        return CanonicalModelBuilder().build(self.import_workbook_bundle(excel_path))

    def _parse_workbook_rows(self, excel_path: str | Path) -> dict[str, Any]:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils.exceptions import InvalidFileException  # type: ignore

        workbook_path = Path(excel_path)
        try:
            wb = load_workbook(workbook_path, data_only=True)
        except Exception as exc:
            rule_id = "invalid-workbook-format" if isinstance(exc, (BadZipFile, InvalidFileException)) else "workbook-open-failed"
            raise ExcelImportValidationError(
                "Workbook could not be opened",
                [
                    ImportDiagnostic(
                        rule_id=rule_id,
                        message="Workbook payload is not a readable .xlsx archive",
                        sheet="(workbook)",
                        value={
                            "path": str(workbook_path),
                            "error_type": type(exc).__name__,
                            "error_message": str(exc),
                        },
                    )
                ],
            ) from exc
        layout_version = self._detect_layout_version(wb.sheetnames)
        if layout_version == "v2-sheeted":
            return self._parse_v2_workbook(wb)
        return self._parse_v1_workbook(wb)

    def _detect_layout_version(self, sheet_names: list[str]) -> str:
        if {"Method", "Assays", "Analytes", "Units"}.issubset(set(sheet_names)):
            return "v2-sheeted"
        return "v1-flat"

    def _parse_v1_workbook(self, workbook: Any) -> dict[str, Any]:
        rows: list[dict[str, Any]] = []
        diagnostics: list[ImportDiagnostic] = []
        for sheet in workbook.worksheets:
            headers, header_map = self._sheet_headers(sheet)
            self._validate_required_columns("v1-flat", sheet.title, headers, diagnostics)
            if diagnostics:
                continue
            mapping = self.COLUMN_MAPPINGS["v1-flat"]["*"]
            seen_rows: set[tuple[str, str, str]] = set()
            for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
                raw_values = [row[header_map[h]].value for h in headers]
                if not any(self._to_string(v) for v in raw_values):
                    continue
                normalized = {src: self._coerce_cell(row[header_map[src]].value) for src in mapping if src in header_map}
                duplicate_key = (
                    self._to_string(normalized.get("AssayKey")),
                    self._to_string(normalized.get("AnalyteKey")),
                    self._to_string(normalized.get("UnitKey")),
                )
                if duplicate_key in seen_rows:
                    diagnostics.append(
                        ImportDiagnostic(
                            rule_id="duplicate-row",
                            message="Duplicate assay/analyte/unit row detected",
                            sheet=sheet.title,
                            row=row_idx,
                            column="AssayKey,AnalyteKey,UnitKey",
                            value={"AssayKey": duplicate_key[0], "AnalyteKey": duplicate_key[1], "UnitKey": duplicate_key[2]},
                        )
                    )
                    continue
                seen_rows.add(duplicate_key)
                rows.append(normalized)

        if diagnostics:
            raise ExcelImportValidationError("Workbook contains validation errors", diagnostics)
        return {"layout_version": "v1-flat", "rows": rows}

    def _parse_v2_workbook(self, workbook: Any) -> dict[str, Any]:
        diagnostics: list[ImportDiagnostic] = []

        method_rows = self._parse_sheet_rows(workbook["Method"], "v2-sheeted", diagnostics)
        assay_rows = self._parse_sheet_rows(workbook["Assays"], "v2-sheeted", diagnostics)
        analyte_rows = self._parse_sheet_rows(workbook["Analytes"], "v2-sheeted", diagnostics)
        unit_rows = self._parse_sheet_rows(workbook["Units"], "v2-sheeted", diagnostics)

        if diagnostics:
            raise ExcelImportValidationError("Workbook contains validation errors", diagnostics)

        return {
            "layout_version": "v2-sheeted",
            "method": method_rows[0] if method_rows else {},
            "assays": assay_rows,
            "analytes": analyte_rows,
            "units": unit_rows,
        }

    def _parse_sheet_rows(self, sheet: Any, layout_version: str, diagnostics: list[ImportDiagnostic]) -> list[dict[str, Any]]:
        headers, header_map = self._sheet_headers(sheet)
        self._validate_required_columns(layout_version, sheet.title, headers, diagnostics)
        if diagnostics:
            return []
        mapping = self.COLUMN_MAPPINGS[layout_version][sheet.title]
        records: list[dict[str, Any]] = []
        seen: set[tuple[Any, ...]] = set()

        mapped_fields = [mapping[h] for h in mapping if h in header_map]
        for row_idx, row in enumerate(sheet.iter_rows(min_row=2), start=2):
            raw_record = {src: row[header_map[src]].value for src in mapping if src in header_map}
            if not any(self._to_string(v) for v in raw_record.values()):
                continue
            record = {mapping[src]: self._coerce_cell(value) for src, value in raw_record.items()}
            duplicate_key = tuple(record.get(field) for field in mapped_fields)
            if duplicate_key in seen:
                diagnostics.append(
                    ImportDiagnostic(
                        rule_id="duplicate-row",
                        message="Duplicate row detected in worksheet",
                        sheet=sheet.title,
                        row=row_idx,
                        column=",".join(mapping.keys()),
                        value=record,
                    )
                )
                continue
            seen.add(duplicate_key)
            records.append(record)
        return records

    def _validate_required_columns(self, layout_version: str, sheet_name: str, headers: list[str], diagnostics: list[ImportDiagnostic]) -> None:
        required = self.REQUIRED_COLUMNS[layout_version][sheet_name if layout_version == "v2-sheeted" else "*"]
        missing = sorted(required - set(headers))
        for column in missing:
            diagnostics.append(
                ImportDiagnostic(
                    rule_id="missing-required-column",
                    message="Required column is missing",
                    sheet=sheet_name,
                    column=column,
                )
            )

    def _sheet_headers(self, sheet: Any) -> tuple[list[str], dict[str, int]]:
        header_row = next(sheet.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            return [], {}
        headers = [self._to_string(c.value) for c in header_row]
        return headers, {name: idx for idx, name in enumerate(headers) if name}


    def _normalize_unit_name(self, value: str) -> str:
        text = self._to_string(value).replace(" ", "")
        aliases = {"ug/ml": "µg/mL", "mg/ml": "mg/mL", "mg/dl": "mg/dL", "mmol/l": "mmol/L"}
        return aliases.get(text.casefold(), text)

    def _split_multi_units(self, value: str) -> list[str]:
        text = self._to_string(value)
        if not text:
            return []
        normalized = text.replace("|", ";").replace(",", ";")
        return [part.strip() for part in normalized.split(";") if part.strip()]

    def _to_string(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return "true" if value else "false"
        if isinstance(value, (int, float)):
            return str(value)
        return str(value).strip()

    def _coerce_cell(self, value: Any) -> str | bool | int | float | None:
        if value is None:
            return None
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value
        text = str(value).strip()
        if text == "":
            return None
        lowered = text.lower()
        if lowered in {"true", "yes", "y", "1"}:
            return True
        if lowered in {"false", "no", "n", "0"}:
            return False
        try:
            if "." in text:
                return float(text)
            return int(text)
        except ValueError:
            return text

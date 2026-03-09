from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

from addon_generator.domain.fragments import FragmentCollection, ProtocolFragment
from addon_generator.domain.issues import IssueSource, IssueSeverity, ValidationIssue, ValidationIssueCollection
from addon_generator.domain.models import AddonModel, AnalyteModel, AnalyteUnitModel, AssayModel, MethodModel, ProtocolContextModel
from addon_generator.generators.analytes_xml_generator import AddonXmlGenerationResult, generate_analytes_addon_xml
from addon_generator.generators.protocol_generator import ProtocolJsonGenerationResult, generate_protocol_json as materialize_protocol_json


@dataclass(slots=True)
class GenerationArtifacts:
    domain_issues: ValidationIssueCollection
    analytes_xml: AddonXmlGenerationResult
    protocol_json: ProtocolJsonGenerationResult


class GenerationService:
    def import_from_excel(self, excel_path: str | Path) -> ProtocolContextModel:
        workbook_path = Path(excel_path)
        try:
            from openpyxl import load_workbook  # type: ignore
        except ModuleNotFoundError as exc:
            raise RuntimeError("openpyxl is required for import_from_excel") from exc

        wb = load_workbook(workbook_path, data_only=True)
        row_payloads: list[dict[str, Any]] = []
        for sheet in wb.worksheets:
            headers = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            for row in sheet.iter_rows(min_row=2):
                values = [cell.value for cell in row]
                if not any(value is not None and str(value).strip() for value in values):
                    continue
                row_payloads.append({headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i]})

        return self.import_from_gui_payload({"rows": row_payloads})

    def import_from_gui_payload(self, payload: dict[str, Any]) -> ProtocolContextModel:
        rows = payload["rows"] if "rows" in payload else [payload]

        methods: dict[str, MethodModel] = {}
        assays: dict[str, AssayModel] = {}
        analytes: dict[str, AnalyteModel] = {}
        units: dict[str, AnalyteUnitModel] = {}

        for row in rows:
            method_info = row.get("MethodInformation", {}) if isinstance(row.get("MethodInformation"), dict) else {}
            assay_info = row.get("AssayInformation", [])
            first_assay = assay_info[0] if isinstance(assay_info, list) and assay_info else {}

            method_name = str(row.get("MethodDisplayName") or method_info.get("DisplayName") or "Method").strip()
            assay_name = str(row.get("AssayDisplayName") or first_assay.get("DisplayName") or "Assay").strip()
            analyte_name = str(row.get("AnalyteName") or "Analyte").strip()
            unit_name = str(row.get("UnitName") or "Unit").strip()

            method_key = f"method:{method_name.casefold()}"
            assay_key = f"assay:{assay_name.casefold()}"
            analyte_key = f"analyte:{assay_key}:{analyte_name.casefold()}"
            unit_key = f"unit:{analyte_key}:{unit_name.casefold()}"

            methods.setdefault(method_key, MethodModel(key=method_key, method_id=len(methods) + 1, display_name=method_name))
            assay_model = assays.setdefault(assay_key, AssayModel(key=assay_key, assay_id=len(assays) + 1, name=assay_name))
            analyte_model = analytes.setdefault(
                analyte_key,
                AnalyteModel(key=analyte_key, analyte_id=len(analytes) + 1, name=analyte_name),
            )
            units.setdefault(unit_key, AnalyteUnitModel(key=unit_key, unit_id=len(units) + 1, name=unit_name, symbol=unit_name))

            if analyte_model not in assay_model.analytes:
                assay_model.analytes.append(analyte_model)
            unit_model = units[unit_key]
            if unit_model not in analyte_model.units:
                analyte_model.units.append(unit_model)

        addon_name = str(payload.get("MethodInformation", {}).get("DisplayName") or payload.get("addon_name") or "Generated Addon")
        addon_id_raw = payload.get("MethodInformation", {}).get("Id") or payload.get("addon_id") or 0
        try:
            addon_id = int(addon_id_raw)
        except (TypeError, ValueError):
            addon_id = 0
        addon = AddonModel(addon_id=addon_id, addon_name=addon_name, methods=list(methods.values()), assays=list(assays.values()))
        return ProtocolContextModel(
            addon=addon,
            method_index=methods,
            assay_index=assays,
            analyte_index=analytes,
            unit_index=units,
        )

    def validate_domain(self, context: ProtocolContextModel) -> ValidationIssueCollection:
        issues = ValidationIssueCollection()
        if not context.addon.addon_name:
            issues.add(
                ValidationIssue(
                    code="missing-addon-name",
                    message="Addon name is required",
                    path="addon.addon_name",
                    severity=IssueSeverity.ERROR,
                    source=IssueSource.DOMAIN,
                )
            )
        if not context.addon.assays:
            issues.add(
                ValidationIssue(
                    code="missing-assays",
                    message="At least one assay is required",
                    path="addon.assays",
                    severity=IssueSeverity.ERROR,
                    source=IssueSource.DOMAIN,
                )
            )
        return issues

    def generate_analytes_xml(
        self, context: ProtocolContextModel, xsd_path: str | Path, output_path: str | Path | None = None
    ) -> AddonXmlGenerationResult:
        return generate_analytes_addon_xml(context.addon, xsd_path=xsd_path, output_path=output_path)

    def generate_protocol_json(self, context: ProtocolContextModel, protocol_fragments: FragmentCollection) -> ProtocolJsonGenerationResult:
        return materialize_protocol_json(context, protocol_fragments)

    def generate_all(
        self,
        context: ProtocolContextModel,
        protocol_fragments: FragmentCollection,
        xsd_path: str | Path,
        xml_output_path: str | Path | None = None,
    ) -> GenerationArtifacts:
        domain_issues = self.validate_domain(context)
        xml_result = self.generate_analytes_xml(context=context, xsd_path=xsd_path, output_path=xml_output_path)
        protocol_result = self.generate_protocol_json(context=context, protocol_fragments=protocol_fragments)
        return GenerationArtifacts(domain_issues=domain_issues, analytes_xml=xml_result, protocol_json=protocol_result)


def fragments_from_protocol_payload(payload: dict[str, Any]) -> FragmentCollection:
    fragments = FragmentCollection()
    for key in ("MethodInformation", "AssayInformation", "LoadingWorkflowSteps", "ProcessingWorkflowSteps"):
        if key in payload:
            fragments.add(ProtocolFragment(path=(key,), value=payload[key], origin="ui"))
    return fragments
